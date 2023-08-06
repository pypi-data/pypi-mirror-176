from time import time

from openpyxl import load_workbook
from xlsxwriter import Workbook
from pyEasyWeb3 import dataTypeClass


class exelData:
    def listNamedTupleToXlsx(data:list[dataTypeClass], fileName:str=None) -> None:
        '''Creates an xlsx file.'''
        data = [namedTuple._asdict() for namedTuple in data]
        workbook = Workbook(rf'{fileName}.xlsx' if fileName else rf'dataQuanty_{len(data)}Time_{int(time())}.xlsx')
        worksheet = workbook.add_worksheet()
        for id,key in enumerate(data[0].keys()):
            worksheet.write(0, id, key)

        for id,accountData in enumerate(data,1):
            for col,dataKey in enumerate(accountData.keys()):
                worksheet.write(id, col, accountData[dataKey])
        workbook.close()

    def readXlsxFile(path:str, columnList:list[str]) -> dict[list]:
        '''Reads columns in an xlsx file.'''
        result = {}
        wb = load_workbook(filename = rf'{path}')

        for columnName in columnList:
            columnData = []
            column = wb.active[columnName]
            for x in range(len(column)):    
                cellData = column[x].value
                if cellData != None:
                    columnData.append(cellData)
            result[columnName] = columnData
        return result