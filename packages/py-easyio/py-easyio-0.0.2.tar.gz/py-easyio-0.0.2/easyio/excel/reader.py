from typing import Any, Generator

from openpyxl import load_workbook
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from easyio.base import EasyReaderBase, string_symbol_split

T_XLSX_READ_ROW = tuple[Any] | dict[str, Any]


class ExcelReader(EasyReaderBase):
    __slots__ = (
        'read_only',
        'data_only',
        'keep_vba',
        'keep_links'
    )

    def __enter__(self):
        return self

    def __exit__(self, exc_type, exc_val, exc_tb):
        if self.read_only:
            self.close()

    def __init__(self,
                 file_path: str,
                 column_config: str = None,
                 read_only: bool = True,
                 data_only: bool = True,
                 keep_vba: Any = None,
                 keep_links: bool = False):
        """
        :param file_path: file path to load
        :param column_config: A key-value map of xlsx-header and field you want.
                              It could be a path to some config file or just a string.
            For example:
                - A xlsx file with two sheets[Sheet1, Sheet2]
                    Sheet1:
                    |Name  |Age  |
                    |Peter |55   |
                    |Brain |24   |
                    Sheet2:
                    |Book  |Price|
                    |Pycode|$56  |
                    |Cook  |$99  |
                - column_config:
                    1. A config file path: [config.yaml]
                        Sheet1:
                            Name: user_name
                            Age: user_age
                        Sheet2:
                            Book: book_name
                            Price: book_price
                        --> Out put after reading:
                        [{'user_name': 'Peter', 'age': 55}...]
                    2. Or, you can just pass a string format
                        'Sheet1:Name::user_name,Age::user_age;Sheet2:Book::book_name,Price::price_name'
            By the way, if there were no header in your xlsx file like:
            - Xlsx without header
                |Peter|55|
                |Brain|66|
            Then you can specify config like this:
            - config.yaml:
                Sheet1:
                    - user_name  # This means the first column will be regarded as 'user_name'
                    - user_age   # And the second column will be 'user_age'
        @Note: Params below are passed directly to openpyxl.load_workbook.
        :param read_only: look at openpyxl.load_workbook
        :param data_only: look at openpyxl.load_workbook
        :param keep_vba: look at openpyxl.load_workbook
        :param keep_links: look at openpyxl.load_workbook
        """
        super().__init__(file_path)
        self._wb = load_workbook(file_path, read_only=read_only, data_only=data_only, keep_vba=keep_vba,
                                 keep_links=keep_links)
        self._sheets = None
        self._sheet_names = None
        self.column_config = column_config
        self.read_only = read_only

    @property
    def wb(self) -> Workbook:
        return self._wb

    def close(self):
        if self.read_only:
            self.wb.close()

    @property
    def sheets(self) -> list[Worksheet]:
        if self._sheets is None:
            self._sheets = self.wb.worksheets
        return self._sheets

    @property
    def sheet_names(self) -> list[str]:
        if self._sheet_names is None:
            self._sheet_names = self.wb.sheetnames
        return self._sheet_names

    def get_sheet_by_name(self, sheet_name: str) -> Worksheet:
        return self.wb[sheet_name]

    def get_sheet(self, sheet_name: str = None) -> Worksheet:
        return self.get_sheet_by_name(sheet_name) if sheet_name is not None else self.sheets[0]

    def get_sheet_by_index(self, index: int) -> Worksheet:
        return self.wb.worksheets[index]

    def iter_sheet_row(self,
                       sheet: Worksheet,
                       min_row: int = None,
                       max_row: int = None,
                       values_only: bool = True,
                       cols: str | list[str] = None,
                       ):
        if cols is None:
            # Return all columns by row
            yield from sheet.iter_rows(values_only=values_only, min_row=min_row, max_row=max_row)
        else:
            # Still by row, but only the specified columns
            if isinstance(cols, str):
                cols = string_symbol_split(cols)
            if not isinstance(cols, (list, tuple)):
                raise TypeError(f"cols should be type list or tuple, but got {type(cols)}")
            col_index_set = set()
            header_row = None
            for col in cols:
                if ':' in col:
                    # like 1:3, means columns from 1 to 2
                    for col_num in self._get_col_range(col):
                        col_index_set.add(col_num)

                elif col.isnumeric():
                    col_index_set.add(int(col) - 1)
                else:
                    # Try to get header(row 1), and get the index of column according to header.
                    if header_row is None:
                        header_row: tuple | None = list(sheet.iter_rows(min_row=1,
                                                                        max_row=1,
                                                                        values_only=values_only))[0]

                    try:
                        col_index = header_row.index(col)

                    except ValueError:
                        raise ValueError(f"Unknown column {col}")
                    col_index_set.add(col_index)

            if any([min_row, max_row]):
                min_row = min_row or 1
                max_row = max_row or sheet.max_row
                row_range = range(min_row - 1, max_row)
            else:
                row_range = range(sheet.max_row)

            for row_num in row_range:
                yield from self._cell_by_cols(row_num, col_index_set, sheet, values_only=values_only)

    @staticmethod
    def _cell_by_cols(
            row_num: int,
            col_index_set: set[int],
            sh: Worksheet,
            values_only: bool = True
    ):

        # Note: row_num and col are 0-based
        cells = (
            sh.cell(row_num + 1, col + 1)
            for col in col_index_set
        )

        if values_only:
            yield tuple(cell.value for cell in cells)
        else:
            yield tuple(cells)

    @staticmethod
    def _get_col_range(col_range: str):
        try:
            lo, hi = filter(lambda x: x.isnumeric(), col_range.split(':'))
        except ValueError as ve:
            raise ValueError(f"Unknown columns range {col_range}."
                             f"If you want to specify column range to read, "
                             f"for example, use '1:3' to get column from 1 to 3."
                             f"Raw error: {ve.args}")
        lo, hi = int(lo) - 1, int(hi) - 1
        if lo > hi:
            raise ValueError(f'Invalid columns range {col_range}. '
                             f'The lower column number {lo} should <= {hi}')
        if not all([lo >= 0, hi > 0]):
            raise ValueError(f'The smaller column number {lo} and '
                             f'the bigger column number {hi} should both greater than 0')
        return range(lo, hi)

    def read_sheet(self, sheet: str = None,
                   values_only: bool = True,
                   cols: list[str] | str = None,
                   min_row: int = None,
                   max_row: int = None) -> T_XLSX_READ_ROW:
        """read rows from some sheet, if sheet is not specified, return the first sheet rows"""
        sh = self.get_sheet(sheet_name=sheet)
        yield from self.iter_sheet_row(sh, values_only=values_only, cols=cols, min_row=min_row, max_row=max_row)

    def read_sheets(self,
                    values_only: bool = True,
                    sheet_cols: dict[str, str | list[str]] = None,
                    ) -> Generator[tuple[str, Generator[T_XLSX_READ_ROW, None, None]], None, None]:
        """
        Read sheets content from xlsx file, by row.
        :param values_only: True, only return cell's value;
                            False return openpyxl.cell.Cell
        :param sheet_cols: {sheet_name: columns of this sheet to read}.
            Columns:
            - 1:3, 5:7 will read columns 1,2,3,5,6,7
            - 1,2,3 will read columns 1,2,3
            - 'Name' will find the index of header row(row 1) and only return that column
        :return:
        """
        for sh in self.sheet_names:
            yield sh, self.read_sheet(sh, values_only=values_only,
                                      cols=sheet_cols.get(sh, None) if sheet_cols is not None else None)

    def read_sheet_as_dict(self, sheet_name: str = None, header: list[str] | dict[str, str] = None):
        """
        Read sheet and return each row as dict
        By default, this method will skip the first header row, and will start from second row
        :param sheet_name: sheet to read
        :param header: An ordered header list or a header map to transfer each row from tuple to a dict.
            If no header is specified, the first row will be regarded as header.
        :return:
        """
        sh = self.get_sheet(sheet_name)
        if sh.max_row < 2:
            raise ValueError(f"Read rows as dict requires the sheet has at lease 2 rows. But got {sh.max_row} row")
        if header is None or isinstance(header, dict):
            header_index_map = {}
            for row_num, row in enumerate(self.iter_sheet_row(sh, values_only=True)):
                if row_num == 0:
                    header_index_map = {
                        col_index: col_value
                        for col_index, col_value in enumerate(row)
                        if col_value
                    } if header is None else {
                        col_index: header[col_value]
                        for col_index, col_value in enumerate(row)
                        if col_value and col_value in header
                    }
                else:
                    yield {
                        header_index_map[col_index]: col_value
                        for col_index, col_value in enumerate(row)
                        if col_index in header_index_map
                    }
        elif isinstance(header, list):
            header_index_map = {
                idx: value
                for idx, value in enumerate(header)
                if value
            }
            for row_num, row in enumerate(self.iter_sheet_row(sh, values_only=True, min_row=2)):
                yield {
                    header_index_map[col_index]: col_value
                    for col_index, col_value in enumerate(row)
                    if col_index in header_index_map
                }
        else:
            raise TypeError(f"Header should be type 'None', "
                            f"or type 'list[str]', "
                            f"or type 'dict[str, Any]', but got {type(header)}")


