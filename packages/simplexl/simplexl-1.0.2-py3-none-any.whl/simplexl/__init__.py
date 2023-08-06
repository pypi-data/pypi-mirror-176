__version__ = '1.0.2'

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


class CreateExcel:

    def __init__(self):

        self.wb = Workbook()
        self.wb.remove(self.wb.active)


    def make_header_columns_with_width(self, header_columns, row_data):
        headers_with_width = []
        try:

            for index, column in enumerate(header_columns):
                max_row_width = max([len(str(row[index])) for row in row_data])
                col_width_max = 40 if max_row_width > 40 else max_row_width
                col_width = len(str(column)) + 2 if col_width_max <= len(str(column)) else col_width_max
                headers_with_width.append(
                    (column, col_width)
                )

        except Exception as err:
            print(err)

        return headers_with_width


    def create_sheet(self, col_data, row_data, sheet_name=None, sheet_index=0):

        sheet_options = {'title': sheet_name, 'index': sheet_index} if sheet_name else {}

        work_sheet = self.wb.create_sheet(**sheet_options)
        self.populate_headers_to_excel(work_sheet, col_data, row_data)
        self.populate_row_data_to_excel(work_sheet, row_data)

    def populate_headers_to_excel(self, work_sheet, col_data, row_data):
        try:
            headers = self.make_header_columns_with_width(
                header_columns=col_data,
                row_data=row_data
            )
            row_num = 1
            for col_num, (col_data, col_width) in enumerate(headers, 1):
                cell = work_sheet.cell(row=row_num, column=col_num)
                cell.value = str(col_data).upper()
                cell.font = Font(name='Calibri', bold=True, color='FFFFFF')
                cell.alignment = Alignment(
                    vertical='center',
                    horizontal='center',
                    wrap_text=False
                )

                cell.fill = PatternFill(
                    start_color='5FABE6',
                    end_color='5FABE6',
                    fill_type='solid'
                )
                column_letter = get_column_letter(col_num)
                column_dimensions = work_sheet.column_dimensions[column_letter]
                column_dimensions.width = col_width
        except Exception as err:
            return err

    def populate_row_data_to_excel(self, work_sheet, row_data):
        try:
            row_num = 1
            for row in row_data:
                row_num += 1
                for col_num in range(1, len(row)+1):
                    cell = work_sheet.cell(row=row_num, column=col_num)
                    cell.value = row[col_num-1]
                    cell.alignment = Alignment(vertical='top', wrap_text=False)
        except Exception as err:
            return err

    def save(self, name="generated-simplexl.xlsx"):
        self.wb.save(name)