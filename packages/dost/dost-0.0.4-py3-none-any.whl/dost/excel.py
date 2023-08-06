"""
Excel Module for dost. This module contains functions for working with excel and spreadsheets

Examples:
    >>> excel.get_row_column_count(df=df)
        (10, 5)
    >>> excel.get_single_cell(df=df,column_name="Column1", cell_number=1)
        "abc"
    >>> excel.excel_create_file(output_folder="C:\\Users\\user\\Desktop", output_filename="test.xlsx", output_sheetname="Sheet1")

This module contains the following functions:

- `authenticate_google_spreadsheet(credential_file_path)`: This creates authentication object for google spreadsheet.
- `get_dataframe_from_google_spreadsheet(auth, spreadsheet_url, sheet_name)`: Get dataframe from google spreadsheet.
- `tabular_data_from_website(website_url, table_number)`: Get tabular data from website.
- `upload_dataframe_to_google_spreadsheet(auth, spreadsheet_url, sheet_name, df)`: Upload dataframe to google spreadsheet.
- `create_file(output_folder, output_filename, output_sheetname)`: Create excel file.
- `to_dataframe(input_filepath, input_sheetname, header)`: Convert excel file to dataframe.
- `get_row_column_count(df)`: Get row and column count of dataframe.
- `dataframe_to_excel(df, output_folder, output_filename, output_sheetname, mode)`: Convert dataframe to excel file.
- `set_single_cell(df, column_name, cell_number, value)`: Set single cell value in excel file.
- `get_single_cell(df, column_name, cell_number, header)`: Get single cell value from excel file.
- `get_all_header_columns(df)`: Get all header columns from excel file.
- `get_all_sheet_names(input_filepath)`: Get all sheet names from excel file.
- `drop_columns(df, cols)`: Drop columns from data frame.
- `clear_sheet(df)`: Clear sheet from excel file.
- `remove_duplicates(df, column_name)`: Remove duplicates from excel file.
- `isNaN(value)`: Check if value is NaN.
- `df_from_list(list_of_lists, column_names)`: Create dataframe from list of lists.
- `df_from_string(df_string, word_delimiter, line_delimiter, column_names)`: Create dataframe from string.
- `df_extract_sub_df(df, row_start, row_end, column_start, column_end)`: Extract sub dataframe from dataframe.
- `set_value_in_df(df, row_number, column_number, value)`: Set value in dataframe.
- `get_value_in_df(df, row_number, column_number)`: Get value from dataframe.
- `df_drop_rows(df, row_start, row_end)`: Drop rows from dataframe.
"""


from typing import Union
import pandas as pd
import os
from pathlib import WindowsPath
from dost.helpers import dostify
from typing import List


@dostify(errors=[])
def authenticate_google_spreadsheet(credential_file_path: Union[str, WindowsPath]) -> object:
    # sourcery skip: raise-specific-error
    """Creates authentication object for google spreadsheet.

    Args:
        credential_file_path (WindowsPath): Credential file path.

    Returns:
        object: Authentication object.

    Examples:
        >>> auth = excel.authenticate_google_spreadsheet(credential_file_path="C:\\Users\\user\\Desktop\\credentials.json")
    """

    # Import Section
    import gspread
    from oauth2client.service_account import ServiceAccountCredentials

    # Code section
    if not credential_file_path:
        raise Exception("credential (json) file path cannot be empty")

    scope = ["https://spreadsheets.google.com/feeds", 'https://www.googleapis.com/auth/spreadsheets',
             "https://www.googleapis.com/auth/drive.file", "https://www.googleapis.com/auth/drive"]

    creds = ServiceAccountCredentials.from_json_keyfilename(
        credential_file_path, scope)

    return gspread.authorize(creds)


@dostify(errors=[])
def get_dataframe_from_google_spreadsheet(auth, spreadsheet_url: str, sheet_name: str = "Sheet1") -> pd.DataFrame:
    # sourcery skip: raise-specific-error
    """ Get dataframe from google spreadsheet

    Args:
        auth (object): Authentication object.
        spreadsheet_url (str): Spreadsheet URL.
        sheet_name (str): Sheet name.

    Returns:
        df (pd.DataFrame): Dataframe object.

    Examples:
        >>> excel.get_dataframe_from_google_spreadsheet(auth=auth,spreadsheet_url="https://docs.google.com/spreadsheets/d/1X2X3X4X5X6X7X8X9X/edit#gid=0", sheet_name="Sheet1")
        df

    """

    # import section
    import pandas as pd

    # Code section
    if not auth:
        raise Exception(
            "Please call authenticate_google_spreadsheet function to get auth")

    if not spreadsheet_url:
        raise Exception("spreadsheet url cannot be empty")

    sh = auth.open_by_url(url=spreadsheet_url)

    # get all the worksheets from sh
    worksheet_list = sh.worksheets()

    # check if sheet_name is already present in worksheet_list
    sheet_present = False
    for worksheet in worksheet_list:
        if worksheet.title == sheet_name:
            sheet_present = True
            break

    if not sheet_present:
        raise Exception("Sheet name not found")
    else:
        worksheet = sh.worksheet(sheet_name)

    return pd.DataFrame(worksheet.get_all_records())


@dostify(errors=[])
def tabular_data_from_website(website_url: str, table_number: int = 1) -> pd.DataFrame:
    # sourcery skip: raise-specific-error
    """Returns a dataframe from a website table.

    Args:
       website_url (str): Website URL.
       table_number (int, optional): Table number. Defaults to 1.

    Examples:
        >>> excel.tabular_data_from_website(website_url="https://en.wikipedia.org/wiki/Wiki")
    """

    # Import Section
    import pandas as pd

    # Code Section
    if not website_url:
        raise Exception("Website URL cannot be empty")

    all_tables = pd.read_html(website_url)

    if not table_number:
        return all_tables
    if table_number > len(all_tables):
        raise Exception(
            "Table number cannot be greater than number of tables")

    if table_number < 1:
        raise Exception("Table number cannot be less than 1")

    return all_tables[table_number - 1]

# @dostify(errors=[])


def upload_dataframe_to_google_spreadsheet(auth, spreadsheet_url: str, sheet_name: str, df: pd.DataFrame) -> None:
    # sourcery skip: raise-specific-error
    """Uploads a dataframe to a google spreadsheet.

    Args:
        auth (object): Authentication object.
        spreadsheet_url (str): Spreadsheet URL.
        sheet_name (str): Sheet name.
        df (pd.DataFrame): Dataframe object.

    Examples:
        >>> excel.upload_dataframe_to_google_spreadsheet(auth=auth, spreadsheet_url="https://docs.google.com/spreadsheets/d/1X2X3X4X5X6X7X8X9X/edit#gid=0", sheet_name="Sheet1", df=df)
    """

    # import section
    from gspread_dataframe import set_with_dataframe
    import pandas as pd

    if not auth:
        raise Exception(
            "Please call authenticate_google_spreadsheet function to get auth")

    if not spreadsheet_url:
        raise Exception("spreadsheet url cannot be empty")

    if not isinstance(df, pd.DataFrame):
        raise Exception("dataframe must be a pandas dataframe")

    sh = auth.open_by_url(url=spreadsheet_url)

    # get all the worksheets from sh
    worksheet_list = sh.worksheets()

    sheet_present = any(worksheet.title ==
                        sheet_name for worksheet in worksheet_list)

    if sheet_present:
        # append df to existing sheet
        worksheet = sh.worksheet(sheet_name)
        row_count = worksheet.get_all_values().__len__()

        if row_count == 0:
            set_with_dataframe(worksheet, dataframe=df)
        else:
            set_with_dataframe(worksheet, dataframe=df,
                               row=row_count + 1, include_column_header=False)

    else:
        worksheet = sh.add_worksheet(
            title=sheet_name, rows="999", cols="26")
        set_with_dataframe(worksheet, df)


@dostify(errors=[])
def create_file(output_folder: Union[str, WindowsPath], output_filename: str, output_sheetname: Union[str, List[str]] = "Sheet1") -> None:
    # sourcery skip: raise-specific-error
    """ Creates an excel file with a sheet in the specified folder.

    Args:
        output_folder (WindowsPath): Output folder path.
        output_filename (str): Output file name.
        output_sheetname (str, optional): Output sheet name. Defaults to "Sheet1".

    Examples:
        >>> excel.create_file(output_folder="C:\\Users\\user\\Desktop", output_filename="test.xlsx", output_sheetname="Sheet1")
    """

    # Import Section
    import os
    from pathlib import Path
    from openpyxl import Workbook

    # Code Section
    output_folder = Path(output_folder)
    if not output_filename:
        raise Exception("Excel File Name cannot be empty")

    if not output_folder:
        raise Exception("Output folder name cannot be empty")
    
    os.makedirs(output_folder,exist_ok=True)

    output_filename = os.path.join(output_folder, f"{str(Path(output_filename).stem)}.xlsx") if ".xlsx" not in output_filename else os.path.join(
        output_folder, output_filename)

    wb = Workbook()
    ws = wb.active
    if (isinstance(output_sheetname, list)):
        ws.title = output_sheetname[0]
        for sheet in output_sheetname[1:]:
            wb.create_sheet(sheet)
    else:
        ws.title = output_sheetname
    wb.save(filename=output_filename)
    wb.close()


@dostify(errors=[])
def _valid_data(input_filepath: Union[str, WindowsPath], input_sheetname: str = "", validate_filepath: bool = True, validate_sheetname: bool = True) -> bool:
    # sourcery skip: raise-specific-error
    """This function validates the input file path and sheet name.

    Args:
        input_filepath (WindowsPath): Input file path.
        input_sheetname (str): Input sheet name.
        validate_filepath (bool, optional): Whether to validate file path or not. Defaults to True.
        validate_sheetname (bool, optional): Whether to validate sheet name or not. Defaults to True.

    Returns:
        bool: True if valid, False if invalid.

    Examples:
        >>> excel._valid_data(input_filepath="C:\\Users\\user\\Desktop\\test.xlsx", input_sheetname="Sheet1")
        True
    """
    # Import Section
    import os
    from openpyxl import load_workbook

    if validate_filepath:
        # Code Section
        input_filepath = str(input_filepath)
        if ".xlsx" not in input_filepath:
            raise Exception(
                "Please provide the excel file name with .xlsx extension")
        if not os.path.exists(input_filepath):
            raise Exception(
                "Please provide the excel file name with correct path")
        if validate_sheetname:
            wb = load_workbook(input_filepath)
            sheet_names = wb.sheetnames
            input_sheetname = input_sheetname
            if input_sheetname not in sheet_names:
                raise Exception(
                    "Please provide the correct sheet name")
        return True


@dostify(errors=[(ValueError, '')])
def to_dataframe(input_filepath: Union[str, WindowsPath], input_sheetname: str, header: int = 1) -> pd.DataFrame:
    # sourcery skip: raise-specific-error
    """Converts excel file to dataframe.

    Args:
        input_filepath (Union[str,WindowsPath]): Input file path.
        input_sheetname (str): Input sheet name.
        header (int, optional): Header row number. Defaults to 1.

    Returns:
        pd.DataFrame: Dataframe of the excel file.

    Examples:
        >>> excel.to_dataframe(input_filepath="C:\\Users\\user\\Desktop\\test.xlsx", input_sheetname="Sheet1")
        dataframe
    """

    # Import Section
    import pandas as pd

    # Code Section
    if not input_filepath:
        raise Exception("Please provide the excel path")
    if not input_sheetname:
        raise Exception("Please provide the sheet name")

    if not _valid_data(input_filepath, input_sheetname):
        raise ValueError("File does not contain valid data")
    data = pd.DataFrame()
    if header > 0:
        data = pd.read_excel(
            input_filepath, sheet_name=input_sheetname, header=header-1, engine='openpyxl')
    elif header == 0:
        data = pd.read_excel(
            input_filepath, sheet_name=input_sheetname, header=None, engine='openpyxl')
    else:
        ValueError('Header value cannot be negative')
    return data


@dostify(errors=[])
def get_row_column_count(df: pd.DataFrame) -> tuple:
    # sourcery skip: raise-specific-error
    """ Returns the row and column count of the dataframe

    Args:
        df (pandas dataframe): Dataframe of the excel file.

    Returns:
        tuple: Row and column count of the dataframe.

    Examples:
        >>> excel.get_row_column_count(df=df)
        (10, 5)
    """

    # Import Section
    import pandas as pd

    # Code Section
    if not isinstance(df, pd.DataFrame):
        raise Exception("Please provide the dataframe")
    return df.shape


@dostify(errors=[])
def dataframe_to_excel(df: pd.DataFrame, output_folder: Union[str, WindowsPath], output_filename: str, output_sheetname: str = "Sheet1", mode: str = 'a') -> None:
    # sourcery skip: raise-specific-error
    """ Converts the dataframe to excel file

    Args:
        df (pandas dataframe): Dataframe of the excel file.
        output_folder (WindowsPath): Output folder path.
        output_filename (str): Output file name.
        output_sheetname (str, optional): Output sheet name. Defaults to "Sheet1".
        mode (str, optional): Mode of the excel file. Defaults to 'a'.

    Examples:
        >>> excel.dataframe_to_excel(df=df, output_folder="C:\\Users\\user\\Desktop", output_filename="test.xlsx", output_sheetname="Sheet1", mode='a')
    """

    # import section
    import pandas as pd
    import os
    from pathlib import Path

    output_folder = Path(output_folder)
    # Code Section
    if not output_folder:
        raise Exception("Output Folder name cannot be empty")

    if not output_filename:
        output_filename = "file"
    
    os.makedirs(output_folder,exist_ok=True)

    output_filepath = os.path.join(output_folder, f"{str(Path(output_filename).stem)}.xlsx") if ".xlsx" not in output_filename else os.path.join(
        output_folder, output_filename)

    if not output_sheetname:
        raise Exception("Please provide the sheet name")

    if not isinstance(df, pd.DataFrame):
        raise Exception("Please provide the dataframe")

    new_file = not os.path.exists(output_filepath)
    if mode == 'a' and not new_file:
        with pd.ExcelWriter(output_filepath, mode="a", engine="openpyxl", if_sheet_exists="overlay",) as writer:
            current_df = to_dataframe(
                output_filepath, output_sheetname)[1]
            row_count = get_row_column_count(current_df)[1]
            df.to_excel(writer, sheet_name=output_sheetname,
                        index=False, startrow=int(row_count[0]), header=False)
    else:
        with pd.ExcelWriter(output_filepath, engine="openpyxl",) as writer:
            df.to_excel(writer, sheet_name=output_sheetname, index=False)


@dostify(errors=[])
def set_single_cell(df: pd.DataFrame, column_name: str, cell_number: int, value: str) -> pd.DataFrame:
    # sourcery skip: raise-specific-error
    """
    Description:
        Writes the given text to the desired column/cell number for the given excel file
    Args:
        df (pandas dataframe): Dataframe of the excel file.
        column_name (str, optional): Column name of the excel file. Defaults to "".
        cell_number (int, optional): Cell number of the excel file. Defaults to 1.
        value (str, optional): Text to be written to the excel file. Defaults to "".

    Returns:
        data (df): Modified dataframe

    Examples:
        >>> df=excel.set_single_cell(df=df, column_name="Column 1",cell_number= 1, value="abc")
        df
    """

    # import section
    import pandas as pd

    # Code Section
    if not isinstance(df, pd.DataFrame):
        raise Exception("Please provide the dataframe")

    if not column_name:
        raise Exception("Please provide the column name")

    if not value:
        raise Exception("Please provide the value to be set")

    if cell_number < 1:
        raise Exception("Please provide the valid cell number")

    df.at[cell_number-1, column_name] = value
    return df


@dostify(errors=[])
def get_single_cell(df: pd.DataFrame, column_name: str, cell_number: int, header: int = 1) -> str:
    # sourcery skip: raise-specific-error
    """Gets the text from the desired column/cell number for the given excel file

    Args:
        df (pandas dataframe): Dataframe of the excel file.
        column_name (str, optional): Column name of the excel file. Defaults to "".
        cell_number (int, optional): Cell number of the excel file. Defaults to 1.
        header (int, optional): Header row number. Defaults to 1.

    Returns:
        data (str): Text from the desired column/cell number for the given excel file

    Examples:
        >>> excel.get_single_cell(df=df, column_name="Column 1",cell_number= 1)
        "abc"
    """

    # Import Section
    import pandas as pd

    # Code Section
    if not isinstance(df, pd.DataFrame):
        raise Exception("Please provide the dataframe")

    if not column_name:
        raise Exception("Please provide the column name")

    if not isinstance(column_name, list):
        column_name = [column_name]

    if cell_number < 1:
        raise Exception("Please provide the valid cell number")

    data = df.at[cell_number-header-1, column_name[0]]

    return str(data)


@dostify(errors=[])
def get_all_header_columns(df: pd.DataFrame) -> Union[List[str], List[int]]:
    # sourcery skip: raise-specific-error
    """Gets all header columns from the excel file

    Args:
        df (pandas dataframe): Dataframe of the excel file.

    Returns:
        data (list): List of header columns

    Examples:
        >>> excel.get_all_header_columns(df=df)
        ["Column1", "Column2"]
    """

    # import section
    import pandas as pd

    # Code Section
    if not isinstance(df, pd.DataFrame):
        raise Exception("Please provide the dataframe")

    return df.columns.values.tolist()


@dostify(errors=[])
def get_all_sheet_names(input_filepath: Union[str, WindowsPath]) -> List[str]:
    # sourcery skip: raise-specific-error
    """Gets the sheet names from the excel file

    Args:
        input_filepath (str): Path of the excel file.

    Returns:
        data (list): List of sheet names

    Examples:
        >>> excel.get_all_sheet_names(input_filepath="demo")
        ["Sheet1", "Sheet2"]
    """

    # Import Section
    from openpyxl import load_workbook

    # Code Section
    if not input_filepath:
        raise Exception("Please provide the excel path")
    if not _valid_data(input_filepath, validate_sheetname=False):
        raise Exception("Please provide the valid excel path")

    wb = load_workbook(input_filepath)
    return wb.sheetnames


@dostify(errors=[(KeyError, "Please provide valid column names")])
def drop_columns(df: pd.DataFrame, cols: Union[str, List[str]]) -> pd.DataFrame:
    # sourcery skip: raise-specific-error
    """Drops the columns from the excel file

    Args:
        df (pandas dataframe): Dataframe of the excel file.
        cols (str, list(str)): Column name to be dropped.

    Returns:
        data (df): Modified dataframe

    Examples:
        >>> excel.drop_columns(df=df, cols="column_name")
        df
    """

    # Import Section
    import pandas as pd

    # Code Section
    if not isinstance(df, pd.DataFrame):
        raise Exception("Please provide the dataframe")

    if not cols:
        raise Exception(
            "Please provide the column name to be dropped.")

    if not isinstance(cols, list):
        cols = [cols]

    df.drop(cols, axis=1, inplace=True)
    return df


@dostify(errors=[])
def clear_sheet(df: pd.DataFrame) -> pd.DataFrame:
    # sourcery skip: raise-specific-error
    """Clears the sheet

    Args:
        df (pandas dataframe): Dataframe of the excel file.

    Returns:
        data (df): Modified dataframe

    Examples:
        >>> excel.clear_sheet(df=df)
        df
    """

    # Import Section
    import pandas as pd

    # Code Section
    if not isinstance(df, pd.DataFrame):
        raise Exception("Please provide the dataframe")

    # Clears the contents of the sheet
    df.drop(df.index, inplace=True)

    return df


@dostify(errors=[])
def remove_duplicates(df: pd.DataFrame, column_name: Union[str, List[str], int, List[int]]) -> pd.DataFrame:
    # sourcery skip: raise-specific-error
    """Removes the duplicates from the given column

    Args:
        df (pandas dataframe): Dataframe of the excel file.
        column_name (str, optional): Column name of the excel file. Defaults to "".

    Returns:
        data (df): Modified dataframe

    Examples:
        >>> excel.remove_duplicates(df=df, column_name="column 1")
        df
    """

    # Import Section
    import pandas as pd

    # Code Section
    which_one_to_keep = "first"

    if not isinstance(df, pd.DataFrame):
        raise Exception("Please provide the dataframe")

    if not column_name:
        df.drop_duplicates(keep=which_one_to_keep, inplace=True)

    else:
        if not isinstance(column_name, list):
            column_name = [column_name]
        df.drop_duplicates(subset=column_name,
                           keep=which_one_to_keep, inplace=True)

    return df


@dostify(errors=[(ValueError, "Give a valid value")])
def isNaN(value: str) -> bool:
    """Checks if the value is NaN

    Args:
        value (str of number): value to be checked

    Returns:
        bool: True if value is NaN, False otherwise

    Examples:
        >>> excel.isNaN(value="abc")
        False
    """

    # Code Section
    if not value:
        raise ValueError
    import math
    return math.isnan(float(value))


@dostify(errors=[])
def df_from_list(list_of_lists: list, column_names: List[str]) -> pd.DataFrame:
    # sourcery skip: raise-specific-error
    """Converts list of lists to dataframe

    Args:
        list_of_lists (list): list of lists to be converted to dataframe
        column_names (list): column names

    Returns:
        pandas dataframe: dataframe

    Examples:
        >>> excel.df_from_list(list_of_lists=[[1,2,3],[4,5,6]], column_names=["col1", "col2", "col3"])
        dataframe
        ..   col1  col2  col3
        0     1     2     3
        1     4     5     6
    """

    # Import Section
    import pandas as pd

    # Code Section
    if not isinstance(list_of_lists, list):
        raise Exception("Please pass input as list of lists")

    return pd.DataFrame(list_of_lists) if column_names is None else pd.DataFrame(list_of_lists, columns=column_names)


@dostify(errors=[])
def df_from_string(df_string: str, word_delimiter: str = " ", line_delimiter: str = "\n", column_names: list = None) -> pd.DataFrame:
    # sourcery skip: raise-specific-error
    """Converts string to dataframe

    Args:
        df_string (str): string to be converted to dataframe
        word_delimiter (str): word delimiter.Defaults to space
        line_delimiter (str): line delimiter. Defaults to new line
        column_names (list): column names. Defaults to None

    Returns:
        pandas dataframe: dataframe

    Examples:
        >>> print(excel.df_from_string(df_string="a b c;d e f",word_delimiter=" ",line_delimiter= ";",column_names= ["Column 1","Column 2","Column 3"]))
        dataframe 
        ..   Column 1  Column 2  Column 3
        0     a     b     c
        1     d     e     f

    """

    # Import Section
    import pandas as pd

    # Code Section
    if not df_string:
        raise Exception("Please pass input as string")

    if not isinstance(df_string, str):
        df_string = df_string

    if column_names is None:
        data = pd.DataFrame([x.split(word_delimiter)
                            for x in df_string.split(line_delimiter)])
    elif isinstance(column_names, list):
        data = pd.DataFrame([x.split(word_delimiter) for x in df_string.split(
            line_delimiter)], columns=column_names)
    return data


@dostify(errors=[])
def df_extract_sub_df(df: pd.DataFrame, row_start: int, row_end: int, column_start: int, column_end: int) -> pd.DataFrame:
    # sourcery skip: raise-specific-error
    """Extracts sub dataframe from the given dataframe

    Args:
        df (pandas dataframe): dataframe
        row_start (int): row start (inclusive)
        row_end (int): row end   (exclusive)
        column_start (int): column start (inclusive)
        column_end (int): column end (exclusive)

    Returns:
        pandas dataframe(pandas dataframe): sub dataframe

    Examples:
        >>> excel.df_extract_sub_df(df=df,row_start= 1, row_end=2, column_start=3, column_end=4)
        sub_dataframe

    """

    # Import Section
    import pandas as pd

    # Code Section
    if df.empty:
        raise Exception("Dataframe cannot be empty")

    if isinstance(df, pd.DataFrame):
        data = df.iloc[row_start-1:row_end-1, column_start-1:column_end-1]

    return data


@dostify(errors=[])
def set_value_in_df(df: pd.DataFrame, row_number: int, column_number: int, value: str) -> pd.DataFrame:
    # sourcery skip: raise-specific-error
    """Sets value in dataframe

    Args:
        df (pandas dataframe): dataframe to be modified
        row_number (int): Row number of the cell
        column_number (int): Column number of the cell
        value (str): value to be set in the cell

    Returns:
        pandas dataframe: dataframe with value set

    Examples:
        >>> excel.set_value_in_df(df=df, row_number=1, column_number=2, value="abc")
        modified_dataframe

    """

    # Import Section
    import pandas as pd

    # Code Section
    if df.empty:
        raise Exception("Dataframe cannot be empty")

    if isinstance(df, pd.DataFrame):
        if row_number < 1 or column_number < 1:
            raise Exception(
                "Row and column number should be greater than 0")

        if row_number > df.shape[0] or column_number > df.shape[1]:
            raise Exception(
                "Row and column number should be less than or equal to dataframe shape")

        df.iloc[row_number-1, column_number-1] = value

        return df


@dostify(errors=[])
def get_value_in_df(df: pd.DataFrame, row_number: int, column_number: int) -> str:
    # sourcery skip: raise-specific-error
    """Gets value from dataframe

    Args:
        df (pandas dataframe): dataframe
        row_number (int): Row number of the cell
        column_number (int): Column number of the cell

    Returns:
        str: value in the cell
    Examples:
        >>> excel.get_value_in_df(df=df, row_number=1, column_number=2)
        abc

    """

    # Import Section
    import pandas as pd

    # Code Section
    if df.empty:
        raise Exception("Dataframe cannot be empty")

    if isinstance(df, pd.DataFrame):
        if row_number < 1 or column_number < 1:
            raise Exception(
                "Row and column number should be greater than 0")

        if row_number > df.shape[0] or column_number > df.shape[1]:
            raise Exception(
                "Row and column number should be less than or equal to dataframe shape")

        data = df.iloc[row_number-1, column_number-1]

    return str(data)


@dostify(errors=[])
def df_drop_rows(df: pd.DataFrame, row_start: int, row_end: int) -> pd.DataFrame:
    # sourcery skip: raise-specific-error
    """Drops rows from dataframe

    Args:
        df (pandas dataframe): dataframe
        row_start (int): row start (inclusive)
        row_end (int): row end   (exclusive)

    Returns:
        pandas dataframe: dataframe with rows dropped
    Examples:
        >>> df = excel.df_drop_rows(df=df, row_start=1, row_end=2)

    """

    # Import Section
    import pandas as pd

    # Code section
    if df.empty:
        raise Exception("Dataframe cannot be empty")

    if isinstance(df, pd.DataFrame):
        # -1 because index starts from 0
        data = df.drop(df.index[row_start-1:row_end-1])

    return data
