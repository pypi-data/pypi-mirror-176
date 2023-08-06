# -*- coding:utf8 -*- #
# -----------------------------------------------------------------------------------
# ProjectName:   cmd_tools
# FileName:     excel_tools.py
# Author:      Jakiro
# Datetime:    2022/5/25 10:29
# Description:
# 命名规则  文件名小写字母+下划线，类名大驼峰，方法、变量名小写字母+下划线连接
# 常量大写，变量和常量用名词、方法用动词
# -----------------------------------------------------------------------------------

import openpyxl,os
from cli.common.context_base import context_manager


class ExcelHandle():
    # 获取开启的源数据文件目录
    # 获取目标文件目录
    def __init__(self, file_path=None, sheet_name=None, ignore_first_row=False, index=6) -> None:
        '''
        初始化excel文件的信息
        :param file_path: 传入一个excel文件目录
        :param sheet_name: 如果只需要单sheet页数据 可传入此参数
        :param ignore_first_row: 是否要忽略首行内容
        :param  index： 认为前几列数据需要打印，否则会打印一堆None,默认为打印6列，参数过大执行时间会较长
        '''
        self.ignore_first_row = ignore_first_row
        self.sheet_name = sheet_name
        self.sheet_count_dict = {}
        self.current_sheet_count = 0
        self.index = index
        if file_path:
            # print(file_path)
            if os.path.exists(file_path):
                print('文件路径存在')
                self.wb = openpyxl.load_workbook(file_path, data_only=True)
            else:
                raise ValueError(f'Wrong file path:{file_path}')
        self.all_data_ = {}

    def __del__(self):
        try:
            if self.wb:
                self.wb.close()

        except:
            pass

    def excel_data_to_python(self) -> dict:
        '''
        将excel文件数据转化为字典
        :return: 将excel中的所有sheet页元素以{first_sheet:[[first_cell,second_cell],second_row,...]，second_sheet:second_data_list,,..}得我形式返回
        '''
        sheet_list = self.wb.sheetnames

        all_dict = {}
        for sheet in sheet_list:
            all_dict[sheet] = self.sheet_data_to_python(sheet)
            self.sheet_count_dict[sheet] = self.current_sheet_count
        return all_dict

    def sheet_data_to_python(self, sheet: str) -> list:
        '''
        将sheet页中的数据转换为列表
        :param sheet:
        :return:根据sheet名，将当前sheet页的数据转换为[[first_cell,second],second_row]
        '''
        # 忽略全空的行
        all_lists = []
        one_line_data_list = []
        self.current_sheet_count = 0
        current_ignore_first_row = self.ignore_first_row
        try:
            ws = self.wb[sheet]
        except:
            raise ValueError('Wrong Sheet Name')

        for row_tuple in ws:
            if current_ignore_first_row:
                current_ignore_first_row = False
                continue
            # 认为前几列数据需要打印，否则会打印一堆None
            for cell in list(row_tuple)[0:self.index]:
                # 拼接 value

                one_line_data_list.append(cell.value)
                # 转化成元组
                one_line_data_tuple = tuple(one_line_data_list)
            if row_tuple[0].value:
                self.current_sheet_count += 1
            # 清空临时列表
            one_line_data_list.clear()
            # 将临时元组添加到总列表中
            all_lists.append(one_line_data_tuple)

        return all_lists

    @property
    def all_data(self) -> dict:
        '''
        根据初始化条件，调用自定义的方法，返回符合当前条件的所用数据
        :return:返回符合当前条件的所有数据 {first_sheet:[[first_cell,second_cell],second_row,...]，second_sheet:second_data_list,,..}
        '''
        if self.all_data_:
            return self.all_data_
        else:
            if self.sheet_name:
                self.all_data_[self.sheet_name] = self.sheet_data_to_python(self.sheet_name)
                self.sheet_count_dict[self.sheet_name] = self.current_sheet_count

                return self.all_data_
            else:

                self.all_data_ = self.excel_data_to_python()
                return self.all_data_

    @property
    def sheet_count(self) -> int:
        '''
        :return: 返回当前文件有多少sheet页
        '''
        # 返回有多少sheet页
        return len(self.wb.sheetnames)

    @property
    def count_all_sheet(self) -> dict:
        '''
        :return:返回每页有多少条数据
        '''
        if self.sheet_count_dict:
            return self.sheet_count_dict
        else:
            self.excel_data_to_python()
            return self.sheet_count_dict

    def count_all_sheet_write(self, new_file_path:str) -> None:
        '''
        统计每页数据，并将其保存在新文件中
        :param new_file_path:
        :return:
        '''
        sheet_count_dict = self.count_all_sheet
        sheet_name = sheet_count_dict.keys()
        sheet_count = sheet_count_dict.values()
        row_data_list = zip(sheet_name, sheet_count)
        with self.write_new_file(new_file_path) as new_workbook:
            new_worksheet = new_workbook.create_sheet('页数统计')
            new_worksheet.append(['sheet名', '统计数量'])
            for row_data in row_data_list:
                new_worksheet.append(row_data)


    # 获取下标与sheet页名的关系
    @property
    def sheet_index(self) -> dict:
        sheet_index_ = {}
        sheet_names = self.wb.sheetnames
        for i, sheet_name in enumerate(sheet_names):
            sheet_index_[sheet_name] = i
        return sheet_index_

    @context_manager
    def write_new_file(self, new_file_path: str) -> object:
        '''
        :param new_file_path:
        :return:返回一个进行关闭操作的workbook对象
        '''
        try:
            new_workbook = openpyxl.Workbook()
            yield new_workbook
        finally:
            if len(new_workbook.sheetnames) > 1:
                default_ws = new_workbook['Sheet']
                new_workbook.remove(default_ws)
            new_workbook.save(
                new_file_path)
            new_workbook.close()
            print(f'create {new_file_path} success')

    def copy_sheet(self, new_file_path: str, sheet_index_list: list) -> None:
        '''
        通过传入的下标索引找到符合要求的sheet页，然后拷贝成一个新的excel文件
        :param sheet_index_list:传入一个列表，列表中的元素是对应原excel sheet页的下标索引
        :param new_file_path:新文件的保存路径
        :return:
        '''
        # 粘贴指定的sheet页
        work_sheet_list = self.wb.sheetnames
        # print(work_sheet_list)
        work_sheets = []
        with self.write_new_file(new_file_path) as new_workbook:
            # 拿到符合要求的 sheet_name_key
            for sheet_index in sheet_index_list:
                # print(sheet_index)
                work_sheets.append(work_sheet_list[int(sheet_index)])
            for new_sheet in work_sheets:
                new_worksheet = new_workbook.create_sheet(new_sheet)
                # print(new_sheet)
                data_lists = self.all_data[new_sheet]
                for data_list in data_lists:
                    new_worksheet.append(data_list)
            return new_file_path

    def copy_excel_all_data(self, new_file_path: str, col: int = None, field: list = None) -> None:
        '''
        将原excel文件按照条件筛选，将筛选后的文件保存到新路径
        :param new_file_path:新文件的保存路径
        :param col:要进行筛选的列号 取值范围 int
        :param field:筛选的字段，当文件数据属于这个字段时，认为其满足需求
        :return:
        '''
        # 拿到表格中的所有数据
        all_data = self.all_data
        type_sum_number = 0
        sheet_count_dict = {}
        with self.write_new_file(new_file_path) as new_workbook:
            for sheet in all_data.keys():
                sheet_count = 0
                status = True
                current_sheet_data = all_data[sheet]
                new_worksheet = new_workbook.create_sheet(sheet)
                for row_data_list in current_sheet_data:
                    if not col:
                        new_worksheet.append(row_data_list)
                    else:
                        if status:
                            new_worksheet.append(row_data_list)
                            status = False
                        else:
                            if row_data_list[col] in field:
                                new_worksheet.append(row_data_list)
                                type_sum_number += 1
                                sheet_count += 1

                sheet_count_dict[sheet] = sheet_count
            new_workbook.save(new_file_path)
            sum_all_count = 0
            for v in self.sheet_count_dict.values():
                sum_all_count += v
            sum_all_count -= len(self.sheet_count_dict)

            print('总用例数量为', sum_all_count)
            print('当前类别数量为', type_sum_number)

    def write_data(self, data_list: list, new_file_path: str, first_data_list: list, data_position: list = None,
                   sheet_name: str = 'new_sheet') -> None:
        '''
        将列表中的数据，保存到excel中，暂时只支持保存到一个sheet页
        :param data_list:列表  [[],[],...]
        :param first_data_list:首行字段 [字段1，字段2，...]
        :param data_position:[列表中第一个数据的位置，列表中第二个元素的位置,...]
        :param new_file_path:new_file_path:新文件的保存路径
        :param sheet_name:保存的sheet页命
        :return:
        '''
        data_position_ = []
        if data_position:
            data_position_ = data_position
        else:
            for i in range(0, len(data_list[0])):
                data_position_.append(i)
        with self.write_new_file(new_file_path) as new_workbook:
            new_worksheet = new_workbook.create_sheet(sheet_name)
            new_worksheet.append(first_data_list)
            for row_number, data_row in enumerate(data_list, start=2):
                data_row_col = zip(data_row, data_position_)
                for data_cell_list in data_row_col:
                    new_worksheet.cell(row=row_number, column=data_cell_list[1], value=data_cell_list[0])


if __name__ == '__main__':
    ex = ExcelHandle(r'/Users/qiujie/study/cmd_tools/data/old/NAC测试用例_客户端.xlsx')
    # print(ex.all_data)
    print(ex.sheet_count, ex.count_all_sheet)
    # ex.copy_excel_all_data(r'data3', col=4, field=['p1', 'p2', 'P1', 'P2'])
    print(ex.sheet_index)
